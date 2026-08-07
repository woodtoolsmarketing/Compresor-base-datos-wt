import pandas as pd
import numpy as np
import os
import re
import sqlite3
import json
import unicodedata
from datetime import datetime

# ==========================================
# BASE DE DATOS PARA HISTORIAL
# ==========================================
DB_NAME = "historial_bases_cargadas.db"

def inicializar_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS historial (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            tipo_carga TEXT,
            ruta TEXT,
            registros_encontrados INTEGER
        )
    ''')
    conn.commit()
    conn.close()

def registrar_historial(tipo_carga, ruta, registros):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute('INSERT INTO historial (fecha, tipo_carga, ruta, registros_encontrados) VALUES (?, ?, ?, ?)',
                   (fecha, tipo_carga, ruta, registros))
    conn.commit()
    conn.close()

def obtener_historial():
    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query('SELECT * FROM historial ORDER BY id DESC', conn)
    conn.close()
    return df

# ==========================================
# NORMALIZACIÓN DE TEXTO
# Todo (catálogo y datos) pasa por acá, así "Cañada de Gómez",
# "CANADA DE GOMEZ" y "cañada  de   gomez" son la misma cosa.
# ==========================================
def normalizar(texto):
    if texto is None: return ""
    txt = str(texto)
    txt = unicodedata.normalize('NFKD', txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = txt.upper()
    txt = re.sub(r'[^A-Z0-9]+', ' ', txt)
    return " ".join(txt.split())

# ==========================================
# CATÁLOGO DE ZONAS (lista oficial WoodTools)
#
# "localidades" = pueblos/ciudades que definen la zona (peso fuerte).
# "debiles"     = nombres de provincia o regiones amplias (peso bajo,
#                 porque aparecen en muchas direcciones sin definir zona).
# "alias"       = como se escribe la zona en las planillas.
# ==========================================
ZONAS_CATALOGO = [
    {
        "codigo": "101", "nombre": "ZONA NORTE", "detalle": "Todos",
        "localidades": [],
        "debiles": [],
        "alias": ["ZONA NORTE", "GBA NORTE", "Z NORTE"],
    },
    {
        "codigo": "102", "nombre": "ZONA SUR (AVELLANEDA / QUILMES)",
        "detalle": "Avellaneda, Quilmes, Berazategui, Florencio Varela y las zonas contiguas",
        "localidades": ["Avellaneda", "Quilmes", "Berazategui", "Florencio Varela", "Florecio Varela"],
        "debiles": [],
        "alias": [],
    },
    {
        "codigo": "103", "nombre": "LA PLATA", "detalle": "Todos",
        "localidades": ["La Plata", "Ensenada", "Berisso"],
        "debiles": [],
        "alias": [],
    },
    {
        "codigo": "104", "nombre": "ZONA SUR (LANUS / LOMAS DE ZAMORA)",
        "detalle": "Lanus, Gerli, Lomas de Zamora, Banfield, Monte Grande, Adrogue, Canning y toda esa parte de la zona sur",
        "localidades": ["Lanus", "Gerli", "Lomas de Zamora", "Banfield", "Monte Grande", "Adrogue", "Canning"],
        "debiles": [],
        "alias": [],
    },
    {
        "codigo": "107", "nombre": "ZONA OESTE", "detalle": "Todos",
        "localidades": [],
        "debiles": [],
        "alias": ["ZONA OESTE", "GBA OESTE", "Z OESTE"],
    },
    {
        "codigo": "110", "nombre": "CABA", "detalle": "Todos",
        "localidades": [],
        "debiles": [],
        "alias": ["CABA", "C A B A", "CAPITAL FEDERAL", "CIUDAD AUTONOMA DE BUENOS AIRES", "ZONA CAPITAL",
                  "Villa Devoto"],
    },
    {
        "codigo": "115", "nombre": "GRAL. BELGRANO / SAN VICENTE",
        "detalle": "Gral. Belgrano, Brandsen, San Vicente, Ranchos, S.M. del Monte",
        "localidades": ["Gral. Belgrano", "General Belgrano", "Brandsen", "San Vicente", "Ranchos",
                        "S.M. del Monte", "San Miguel del Monte"],
        "debiles": [],
        "alias": ["RUTA 29"],
    },
    {
        "codigo": "116", "nombre": "NAVARRO / CAÑUELAS / LOBOS",
        "detalle": "Navarro, Las Heras, Cañuelas, Lobos",
        "localidades": ["Navarro", "Cañuelas", "Lobos", "Gral. Las Heras", "General Las Heras", "Las Heras"],
        "debiles": [],
        "alias": [],
    },
    {
        "codigo": "120", "nombre": "RUTA 2",
        "detalle": "Chascomus, Lezama, Dolores, Maipu, Castelli",
        "localidades": ["Chascomus", "Lezama", "Dolores", "Maipu", "Castelli"],
        "debiles": [],
        "alias": ["RUTA 2"],
    },
    {
        "codigo": "121", "nombre": "CAÑADA DE GOMEZ / SANTA FE CAP.",
        "detalle": "Cañada de Gomez: Armstrong, Arteaga, Cañada de Gomez, Carcaraña, Correa | Santa Fe Cap.: Santa Fe Capital",
        "localidades": ["Armstrong", "Arteaga", "Cañada de Gomez", "Carcaraña", "Correa", "Santa Fe Capital"],
        "debiles": ["Santa Fe"],
        "alias": [],
        "subzonas": [
            {"nombre": "CAÑADA DE GOMEZ", "detalle": "Armstrong, Arteaga, Cañada de Gomez, Carcaraña, Correa"},
            {"nombre": "SANTA FE CAP.", "detalle": "Santa Fe Capital"},
        ],
    },
    {
        "codigo": "122", "nombre": "MAR DEL PLATA / COSTA ATLANTICA",
        "detalle": "Mar del Plata, Miramar, Chapadmalal, Santa Teresita, Mar del Tuyu, San Bernardo, Mar de Ajo, "
                   "Pinamar, San Clemente, Santa Clara, Villa Gesell, Ostende, Las Toninas, Aguas Verdes, Valeria del Mar",
        "localidades": ["Mar del Plata", "Miramar", "Chapadmalal", "Santa Teresita", "Mar del Tuyu", "San Bernardo",
                        "Mar de Ajo", "Pinamar", "San Clemente", "San Clemente del Tuyu", "Santa Clara",
                        "Santa Clara del Mar", "Villa Gesell", "Ostende", "Las Toninas", "Aguas Verdes",
                        "Valeria del Mar"],
        "debiles": [],
        "alias": ["COSTA ATLANTICA"],
    },
    {
        "codigo": "124", "nombre": "NECOCHEA / QUEQUEN",
        "detalle": "Necochea, Quequen, Sierra de la Ventana",
        "localidades": ["Necochea", "Quequen", "Sierra de la Ventana"],
        "debiles": [],
        "alias": [],
    },
    {
        "codigo": "126", "nombre": "SAN JUAN / SAN LUIS / MENDOZA",
        "detalle": "San Luis, Villa Mercedes, V. Mackenna, Mendoza, Tunuyan, Tupungato, San Rafael, Gral. Alvear, "
                   "San Juan, Caucete, Pocitos, Rawson, Godoy Cruz, Palmira, Guaymallen, Lujan de Cuyo, Las Heras, "
                   "Maipu, Gral. San Martin, Chimbas, Albardon",
        "localidades": ["Villa Mercedes", "V. Mackenna", "Vicuña Mackenna", "Tunuyan", "Tupungato", "San Rafael",
                        "Gral. Alvear", "General Alvear", "Caucete", "Pocitos", "Rawson", "Godoy Cruz", "Palmira",
                        "Guaymallen", "Lujan de Cuyo", "Las Heras", "Maipu", "Gral. San Martin", "General San Martin",
                        "Chimbas", "Albardon"],
        "debiles": ["San Luis", "Mendoza", "San Juan", "Cuyo"],
        "alias": ["CUYO"],
    },
    {
        "codigo": "130", "nombre": "RUTA 5",
        "detalle": "Arenales, Sta. Rosa, Darregueira, Trenque Lauquen, Rufino, Olavarria, Junin, Bragado, "
                   "C. Casares, Roque Perez",
        "localidades": ["Arenales", "Gral. Arenales", "General Arenales", "Darregueira", "Trenque Lauquen", "Rufino",
                        "Olavarria", "Junin", "Bragado", "C. Casares", "Carlos Casares", "Roque Perez",
                        "Sta. Rosa", "Santa Rosa"],
        "debiles": [],
        "alias": ["RUTA 5"],
    },
    {
        "codigo": "132", "nombre": "BAHIA BLANCA",
        "detalle": "Pringles, Castex, Rivadavia, Cnel. Pringles, Pigüe, Carhue, Gral. Lacha, Sta. Rosa, Loberia, "
                   "Huanguelen, Cnel. Suarez, Gral. Pico",
        "localidades": ["Pringles", "Cnel. Pringles", "Coronel Pringles", "Castex", "Eduardo Castex", "Rivadavia",
                        "Pigüe", "Carhue", "Gral. Lacha", "Loberia", "Huanguelen", "Cnel. Suarez", "Coronel Suarez",
                        "Gral. Pico", "General Pico", "Sta. Rosa", "Santa Rosa"],
        "debiles": ["La Pampa"],
        "alias": ["BAHIA BLANCA", "RUTA 3"],
    },
    {
        "codigo": "136", "nombre": "ENTRE RIOS",
        "detalle": "Colon, Villa Elisa, Ubajay, San Salvador, Concordia, Mocoreta, Juan Pujol, Monte Caseros, "
                   "San Jose, Concepcion del Uruguay, Gualeguaychu, Chajari, Calabacillas, Pucheta, Colonia Libertad, "
                   "Federacion, Paso de los Libres, Santo Tome, Victoria, Diamante, Gualeguay, Ceibas",
        "localidades": ["Colon", "Villa Elisa", "Ubajay", "San Salvador", "Concordia", "Mocoreta", "Juan Pujol",
                        "Monte Caseros", "San Jose", "Concepcion del Uruguay", "Gualeguaychu", "Chajari",
                        "Calabacillas", "Pucheta", "Colonia Libertad", "Federacion", "Paso de los Libres",
                        "Santo Tome", "Victoria", "Diamante", "Gualeguay", "Ceibas"],
        "debiles": ["Entre Rios"],
        "alias": ["ENTRE RIOS"],
    },
    {
        "codigo": "137", "nombre": "CORDOBA",
        "detalle": "Cordoba, Carlos Paz, Cosquin, Fronteras, Devoto, La Falda, C. del Monte, J. Maria, Rio Ceballos, "
                   "V. del Rosario, Rio II, Alta Gracia, Calamuchita, V.G. Belgrano, Oliva, Oncativo, La Carlota, "
                   "Canals, Rivera Indarte, Chapui, Los Boulevares, Ferreyra, Sampacho, Rio Cuarto, V. Tuerto, Elena, "
                   "Almafuerte, Rio III, Costa Sacate, Monte Buey, Bell Ville, Dehesa, Adelia Maria, Cnia. Caroya, "
                   "Dolores, Colazo, Arroyito, Laguna Larga, La Para, Marcos Juarez",
        "localidades": ["Carlos Paz", "Villa Carlos Paz", "Cosquin", "Fronteras", "Devoto", "La Falda",
                        "C. del Monte", "Capilla del Monte", "J. Maria", "Jesus Maria", "Rio Ceballos",
                        "V. del Rosario", "Villa del Rosario", "Rio II", "Rio Segundo", "Alta Gracia", "Calamuchita",
                        "V.G. Belgrano", "Villa General Belgrano", "Oliva", "Oncativo", "La Carlota", "Canals",
                        "Rivera Indarte", "Chapui", "Los Boulevares", "Ferreyra", "Sampacho", "Rio Cuarto",
                        "V. Tuerto", "Venado Tuerto", "Elena", "Almafuerte", "Rio III", "Rio Tercero",
                        "Costa Sacate", "Monte Buey", "Bell Ville", "Dehesa", "Adelia Maria", "Cnia. Caroya",
                        "Colonia Caroya", "Dolores", "Colazo", "Arroyito", "Laguna Larga", "La Para",
                        "Marcos Juarez"],
        "debiles": ["Cordoba"],
        "alias": ["CORDOBA"],
    },
    {
        "codigo": "140", "nombre": "PERGAMINO",
        "detalle": "S.A. de Areco, Cap. Sarmiento, Salto, S.A. de Giles, C. de Areco, Arrecifes, Pergamino, "
                   "N. de la Riestra, Chacabuco, Chivilcoy, 25 de Mayo, Zarate, Campana",
        "localidades": ["S.A. de Areco", "San Antonio de Areco", "Cap. Sarmiento", "Capitan Sarmiento", "Salto",
                        "S.A. de Giles", "San Andres de Giles", "C. de Areco", "Carmen de Areco", "Arrecifes",
                        "Pergamino", "N. de la Riestra", "Norberto de la Riestra", "Chacabuco", "Chivilcoy",
                        "25 de Mayo", "Veinticinco de Mayo", "Zarate", "Campana"],
        "debiles": [],
        "alias": ["RUTA 8"],
    },
    {
        "codigo": "141", "nombre": "LUJAN / PILAR / MERCEDES",
        "detalle": "Pilar, Capilla del Señor, Exalt. de la Cruz, Lujan, Mercedes, Parada Robles, Villa Flandria, "
                   "Jauregui, Open Door, Cortines",
        "localidades": ["Pilar", "Capilla del Señor", "Exalt. de la Cruz", "Exaltacion de la Cruz", "Lujan",
                        "Mercedes", "Parada Robles", "Villa Flandria", "Jauregui", "Open Door", "Cortines"],
        "debiles": [],
        "alias": ["RUTA 7"],
    },
    {
        "codigo": "142", "nombre": "ROSARIO",
        "detalle": "Acebal, Albarellos, Alvarez, Arequito, Arroyo Seco, Baradero, Casilda, Emp. Villa Constitucion, "
                   "Fuentes, Funes, Pujato, Ramallo, Roldan, Rosario, S. Nicolas de los Arroyos, Soldini, Uranga, "
                   "V. Constitucion, V. Ramallo",
        "localidades": ["Acebal", "Albarellos", "Alvarez", "Arequito", "Arroyo Seco", "Baradero", "Casilda",
                        "Emp. Villa Constitucion", "Empalme Villa Constitucion", "Fuentes", "Funes", "Pujato",
                        "Ramallo", "Roldan", "Rosario", "S. Nicolas de los Arroyos", "San Nicolas de los Arroyos",
                        "Soldini", "Uranga", "V. Constitucion", "Villa Constitucion",
                        "V. Ramallo", "Villa Ramallo"],
        "debiles": [],
        "alias": [],
    },
    {
        "codigo": "143", "nombre": "URDINARRAIN / NOGOYA / CORRIENTES",
        "detalle": "Urdinarrain, Victoria, Crespo, Ramirez, Nogoya, Rosario del Tala, Basavilbaso, Corrientes",
        "localidades": ["Urdinarrain", "Victoria", "Crespo", "Ramirez", "Nogoya", "Rosario del Tala", "Basavilbaso"],
        "debiles": ["Corrientes"],
        "alias": [],
    },
    {
        "codigo": "144", "nombre": "CHILLAR / ALBERTI / SUIPACHA",
        "detalle": "Chillar, Gral. O'Brien, Alberti, Moquehua, Gral. Alvear, Gorostiaga, Suipacha, Villa Lia, "
                   "Pedernales, O'Higgins, Coronel Mom, Ascension, Totoras",
        "localidades": ["Chillar", "Gral. O'Brien", "General O'Brien", "Alberti", "Moquehua", "Gral. Alvear",
                        "General Alvear", "Gorostiaga", "Suipacha", "Villa Lia", "Pedernales", "O'Higgins",
                        "Coronel Mom", "Ascension", "Totoras"],
        "debiles": [],
        "alias": [],
    },
    {
        "codigo": "146", "nombre": "SALTA",
        "detalle": "Salta, Catamarca, Tucuman, Jujuy, La Rioja, Santiago del Estero",
        "localidades": [],
        "debiles": ["Salta", "Catamarca", "Tucuman", "Jujuy", "La Rioja", "Santiago del Estero"],
        "alias": ["NOROESTE", "NOA"],
    },
    {
        "codigo": "147", "nombre": "SUR",
        "detalle": "Tierra del Fuego, Chubut, Santa Cruz, Pto. Madryn, Patagonia Sur, Rio Gallegos, Rio Grande",
        "localidades": ["Pto. Madryn", "Puerto Madryn", "Rio Gallegos", "Rio Grande", "Ushuaia", "Comodoro Rivadavia"],
        "debiles": ["Tierra del Fuego", "Chubut", "Santa Cruz", "Patagonia Sur"],
        "alias": ["PATAGONIA SUR", "SUR II"],
    },
    {
        "codigo": "148", "nombre": "SUR CORTA",
        "detalle": "Rio Colorado, Choele Choel, V. Regina, Lamarque, Gral. Roca, Allen, Cipolletti, 5 Saltos, "
                   "Gral. Conesa, Neuquen, Gral. Godoy, Luis Beltran, Centenario, Plottier, Fernandez Oro",
        "localidades": ["Rio Colorado", "Choele Choel", "V. Regina", "Villa Regina", "Lamarque", "Gral. Roca",
                        "General Roca", "Allen", "Cipolletti", "5 Saltos", "Cinco Saltos", "Gral. Conesa",
                        "General Conesa", "Gral. Godoy", "General Godoy", "Luis Beltran", "Centenario", "Plottier",
                        "Fernandez Oro"],
        "debiles": ["Neuquen", "Rio Negro", "Valle Rio Negro"],
        "alias": ["SUR CORTA", "VALLE RIO NEGRO"],
    },
    {
        "codigo": "149", "nombre": "ESPERANZA",
        "detalle": "Esperanza, San Jeronimo, Franck, Humboldt, San Justo, Rafaela, Crespo, San Francisco (Cba.)",
        "localidades": ["Esperanza", "San Jeronimo", "Franck", "Humboldt", "San Justo", "Rafaela", "San Francisco",
                        "Crespo"],
        "debiles": [],
        "alias": [],
    },
    {
        "codigo": "150", "nombre": "LITORAL / MISIONES / CORRIENTES",
        "detalle": "Litoral, Misiones, Corrientes",
        "localidades": ["Posadas", "Obera", "Eldorado", "Puerto Iguazu"],
        "debiles": ["Litoral", "Misiones", "Corrientes"],
        "alias": ["LITORAL", "MESOPOTAMIA"],
    },
    {
        "codigo": "151", "nombre": "CHACO / FORMOSA",
        "detalle": "Chaco, Formosa",
        "localidades": ["Resistencia", "Saenz Peña", "Presidencia Roque Saenz Peña", "Clorinda"],
        "debiles": ["Chaco", "Formosa"],
        "alias": [],
    },
    {
        "codigo": "152", "nombre": "SUR LARGA",
        "detalle": "Zapala, Alumine, Junin de los Andes, S.M. de los Andes, V. La Angostura, El Bolson, Lago Puelo, "
                   "Esquel, El Maiten, Cutral Co",
        "localidades": ["Zapala", "Alumine", "Junin de los Andes", "S.M. de los Andes", "San Martin de los Andes",
                        "V. La Angostura", "Villa La Angostura", "El Bolson", "Lago Puelo", "Esquel", "El Maiten",
                        "Cutral Co", "Cutral-Co"],
        "debiles": [],
        "alias": ["SUR LARGA"],
    },
]

# Zonas que NO están en la lista oficial pero que pueden aparecer en bases viejas.
# Se reconocen para no perder el dato, pero no se ofrecen como zona asignable.
ZONAS_ESPECIALES = {
    "301": "EXTERIOR",
}

# Códigos viejos que desaparecieron de la lista nueva -> a qué código equivalen hoy.
# Sólo se aplican a códigos que YA NO EXISTEN en el catálogo actual, así nunca
# pisan el significado nuevo de un código que sigue vigente.
ZONAS_LEGADO = {
    "155": "148",   # VALLE RIO NEGRO / NEUQUEN  -> 148 SUR CORTA
    "156": "122",   # MAR DEL PLATA              -> 122 MAR DEL PLATA / COSTA ATLANTICA
}

# Compatibilidad hacia atrás: el resto del programa sigue usando MAPA_ZONAS.
MAPA_ZONAS = {z["codigo"]: z["nombre"] for z in ZONAS_CATALOGO}
MAPA_ZONAS_DETALLE = {z["codigo"]: z["detalle"] for z in ZONAS_CATALOGO}
CODIGOS_ZONA_VALIDOS = set(MAPA_ZONAS.keys())

# --- Índices de búsqueda (se arman una sola vez al importar el módulo) ---
PESO_LOCALIDAD = 3
PESO_ALIAS = 2
PESO_DEBIL = 2
PESO_AMBIGUO = 1

def _construir_indice_zonas():
    """frase_normalizada -> {'codigos': [...], 'peso': n}. Las frases que caen en
    dos zonas distintas (ej: LAS HERAS en 116 y 126) quedan marcadas como ambiguas."""
    indice = {}
    def _sumar(frase, codigo, peso_base):
        clave = normalizar(frase)
        if not clave or len(clave) < 3: return
        entrada = indice.setdefault(clave, {"codigos": [], "peso": peso_base})
        if codigo not in entrada["codigos"]:
            entrada["codigos"].append(codigo)
        entrada["peso"] = min(entrada["peso"], peso_base)

    for zona in ZONAS_CATALOGO:
        for loc in zona.get("localidades", []):
            _sumar(loc, zona["codigo"], PESO_LOCALIDAD)
        for deb in zona.get("debiles", []):
            _sumar(deb, zona["codigo"], PESO_DEBIL)
        for ali in zona.get("alias", []):
            _sumar(ali, zona["codigo"], PESO_ALIAS)

    for entrada in indice.values():
        if len(entrada["codigos"]) > 1:
            entrada["peso"] = PESO_AMBIGUO
    return indice

_INDICE_ZONAS = _construir_indice_zonas()
# Frases ordenadas de la más larga a la más corta: así "JUNIN DE LOS ANDES" (152)
# se consume antes que "JUNIN" (130) y no se cuentan las dos.
_FRASES_ZONA_ORDENADAS = sorted(_INDICE_ZONAS.keys(), key=len, reverse=True)
_PATRONES_ZONA = {f: re.compile(r'(?<![A-Z0-9])' + re.escape(f) + r'(?![A-Z0-9])') for f in _FRASES_ZONA_ORDENADAS}

# Todas las localidades juntas: sirve para "tapar" la geografía antes de buscar
# nombres de vendedores (así SAN LUIS no se confunde con el vendedor LUIS).
_PATRON_GEOGRAFIA = re.compile(
    r'(?<![A-Z0-9])(?:' + "|".join(re.escape(f) for f in _FRASES_ZONA_ORDENADAS) + r')(?![A-Z0-9])'
) if _FRASES_ZONA_ORDENADAS else None

# Un número de 3 cifras suelto, que no sea parte de un número más largo (teléfono, CUIT, importe).
_PATRON_CODIGO_ZONA = re.compile(r'(?<![0-9])([1-3][0-9]{2})(?![0-9])')

def listar_zonas_para_ui():
    """[(codigo, nombre, detalle)] en el orden del catálogo, para las ventanas de configuración."""
    return [(z["codigo"], z["nombre"], z["detalle"]) for z in ZONAS_CATALOGO]

# ==========================================
# CATÁLOGO DE VENDEDORES
#
# "codigos" son TODAS las formas en que ese vendedor aparece escrito en las
# planillas del ERP (con y sin cero adelante). Todas se normalizan al código
# canónico, que es el que se exporta y el que busca el bot de WhatsApp.
# ==========================================
VENDEDORES_CATALOGO = {
    "0":  {"nombre": "VALENTIN / CARLOS", "nombres": ["VALENTIN"],                "codigos": ["0", "00", "000"]},
    "1":  {"nombre": "EMMANUEL",          "nombres": ["EMMANUEL", "EMANUEL", "EMMA"], "codigos": ["1", "01", "001", "302"]},
    "03": {"nombre": "LUIS",              "nombres": ["LUIS"],                    "codigos": ["3", "03", "003"]},
    "05": {"nombre": "ROBERTO",           "nombres": ["ROBERTO", "ROBER"],        "codigos": ["5", "05", "005"]},
    "09": {"nombre": "EZEQUIEL",          "nombres": ["EZEQUIEL", "EZE"],         "codigos": ["9", "09", "009"]},
    "16": {"nombre": "LUCAS",             "nombres": ["LUCAS"],                   "codigos": ["16", "016"]},
    "18": {"nombre": "JORGE",             "nombres": ["JORGE"],                   "codigos": ["18", "018"]},
    "40": {"nombre": "NICOLAS",           "nombres": ["NICOLAS", "NICO"],         "codigos": ["40", "040", "15", "015"]},
    "44": {"nombre": "ALAN",              "nombres": ["ALAN"],                    "codigos": ["44", "044", "4", "04", "004"]},
}

CODIGO_SIN_ASIGNAR = "0"

# "CARLOS" es también Carlos Paz / Carlos Casares, así que no se busca como
# nombre suelto: el vendedor 0 se detecta por código o por VALENTIN.
_ALIAS_CODIGO_VEND = {}
_ALIAS_NOMBRE_VEND = {}
for _canon, _info in VENDEDORES_CATALOGO.items():
    for _c in _info["codigos"]:
        _ALIAS_CODIGO_VEND[normalizar(_c)] = _canon
    for _n in _info["nombres"]:
        _ALIAS_NOMBRE_VEND[normalizar(_n)] = _canon

_NOMBRES_VEND_ORDENADOS = sorted(_ALIAS_NOMBRE_VEND.keys(), key=len, reverse=True)
_PATRONES_NOMBRE_VEND = {n: re.compile(r'(?<![A-Z0-9])' + re.escape(n) + r'(?![A-Z0-9])')
                         for n in _NOMBRES_VEND_ORDENADOS}

# Palabras que, si están justo antes del nombre, indican que es un lugar y no una persona.
_PREFIJOS_GEOGRAFICOS = {
    "SAN", "SANTA", "SANTO", "STA", "STO", "VILLA", "VILA", "GRAL", "GENERAL", "CNEL", "CORONEL",
    "PUERTO", "PTO", "LAGO", "LAGUNA", "SIERRA", "SIERRAS", "COLONIA", "CNIA", "BAHIA", "ARROYO",
    "CAMPO", "PASO", "BARRIO", "CALLE", "AV", "AVDA", "AVENIDA", "RUTA", "PARAJE", "ESTACION",
    "PLAZA", "PARQUE", "PJE", "PASAJE", "DON", "PRESIDENTE", "PTE", "DR", "ING", "CAP", "CAPITAN",
    "ALTE", "ALMIRANTE", "TTE", "TENIENTE", "SGTO", "SARGENTO", "MONTE", "CERRO", "RIO", "MAR",
    "ISLA", "COSTA", "LOMAS", "LOMA", "FRAY", "PADRE", "CIUDAD", "PUEBLO",
}

# Palabras que, si están justo después del nombre, también lo vuelven un lugar.
_SUFIJOS_GEOGRAFICOS = {
    "LUIS": {"BELTRAN", "GUILLON", "PIEDRABUENA", "MARIA", "CHICO"},
    "JORGE": {"NEWBERY", "CHAVEZ"},
    "LUCAS": {"GONZALEZ", "MONTEVERDE"},
    "NICOLAS": {"DE", "DEL"},
    "ROBERTO": {"PAYRO"},
}

def normalizar_codigo_vendedor(valor):
    """Devuelve el código canónico ('44', '40', '09'...) o None si no se reconoce.
    Acepta compuestos tipo '40/15' o '302/1' quedándose con el primer código válido."""
    if valor is None: return None
    crudo = str(valor).strip()
    if not crudo or crudo.lower() in ("nan", "none"): return None

    directo = _ALIAS_CODIGO_VEND.get(normalizar(crudo))
    if directo is not None: return directo

    # Compuestos: 40/15, 302/1, 4-44, "44 y 04"
    partes = re.split(r'[^0-9]+', crudo)
    for parte in partes:
        if not parte: continue
        canon = _ALIAS_CODIGO_VEND.get(normalizar(parte))
        if canon is not None: return canon
    return None

_PATRON_FECHA = re.compile(r'^\s*\d{1,4}[/\-\.]\d{1,2}[/\-\.]\d{1,4}')

def _parece_fecha(valor):
    """Evita que un '15/03/2026' en la columna Vendedor se lea como el código 15."""
    return bool(_PATRON_FECHA.match(str(valor)))

def _celda_es_solo_codigo(valor):
    """True si la celda es únicamente un código (o compuesto de códigos), sin texto alrededor."""
    crudo = str(valor).strip()
    if not crudo or _parece_fecha(crudo): return False
    return bool(re.fullmatch(r'[0-9]{1,3}(?:\s*[/\-y]\s*[0-9]{1,3})*', crudo, flags=re.IGNORECASE))

def _buscar_nombre_vendedor(texto_norm):
    """Busca nombres de vendedores en un texto ya normalizado, esquivando nombres de lugares."""
    for nombre in _NOMBRES_VEND_ORDENADOS:
        patron = _PATRONES_NOMBRE_VEND[nombre]
        for m in patron.finditer(texto_norm):
            previo = texto_norm[:m.start()].split()
            if previo and previo[-1] in _PREFIJOS_GEOGRAFICOS:
                continue
            siguiente = texto_norm[m.end():].split()
            bloqueo = _SUFIJOS_GEOGRAFICOS.get(nombre, set())
            if siguiente and siguiente[0] in bloqueo:
                continue
            return _ALIAS_NOMBRE_VEND[nombre]
    return None

def _tapar_geografia(texto_norm):
    """Reemplaza por espacios todo lo que sea nombre de localidad/zona conocida."""
    if not _PATRON_GEOGRAFIA: return texto_norm
    return _PATRON_GEOGRAFIA.sub(lambda m: " " * len(m.group(0)), texto_norm)

# ==========================================
# CONFIGURACIÓN DE CELULARES Y VÍNCULOS ZONA-VENDEDOR
# ==========================================
ARCHIVO_VEND = "vendedores_config.json"
ARCHIVO_VINCULOS = "vinculos_zonas.json"

def cargar_mapa_vendedores():
    defaults = {
        "0": "1145394279 o 1165630406",
        "1": "1157528428",
        "01": "1157528428",
        "302": "1157528428",
        "40": "1157528427",
        "15": "1157528427",
        "18": "1145640940",
        "16": "1145640831",
        "44": "1156321012",
        "4": "1156321012",
        "04": "1156321012",
        "9": "1153455274",
        "09": "1153455274",
        "3": "1168457778",
        "03": "1168457778",
        "5": "1164591316",
        "05": "1164591316"
    }
    if not os.path.exists(ARCHIVO_VEND):
        try:
            with open(ARCHIVO_VEND, 'w', encoding='utf-8') as f:
                json.dump(defaults, f, indent=4)
        except: pass
        return defaults
    else:
        try:
            with open(ARCHIVO_VEND, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return defaults

def guardar_mapa_vendedores(nuevo_mapa):
    try:
        with open(ARCHIVO_VEND, 'w', encoding='utf-8') as f:
            json.dump(nuevo_mapa, f, indent=4)
        return True
    except:
        return False

# Códigos de zona que cambiaron de significado con la lista nueva.
# clave = código nuevo, valor = código viejo del que hereda la asignación.
_HERENCIA_ZONAS = {
    "141": "124",   # LUJAN / PILAR / MERCEDES  (antes 124 = LUJAN / MERCEDES)
    "122": "156",   # MAR DEL PLATA             (antes 156)
    "148": "155",   # SUR CORTA                 (antes 155 = VALLE RIO NEGRO / NEUQUEN)
    "147": "122",   # SUR                       (antes 122 = SUR II / PATAGONIA)
    "126": "150",   # SAN JUAN / SAN LUIS / MZA (antes 150 = CUYO)
    "150": "152",   # LITORAL / MISIONES / CTES (antes 152 = MISIONES / CORRIENTES)
    "124": None,    # NECOCHEA / QUEQUEN        (zona nueva, no hereda nada)
    "152": None,    # SUR LARGA                 (zona nueva)
    "121": None,    # CAÑADA DE GOMEZ / STA FE  (zona nueva)
    "116": None,    # NAVARRO / CAÑUELAS / LOBOS(zona nueva)
}

def _migrar_vinculos(mapa_viejo):
    """Adapta el vinculos_zonas.json a la lista de zonas nueva:
    - normaliza los códigos de vendedor (04 -> 44, 15 -> 40, etc.)
    - las zonas que cambiaron de significado heredan del código viejo equivalente
    - tira las zonas que ya no existen y agrega las nuevas en 0 (sin asignar)."""
    limpio = {}
    for cod, vend in (mapa_viejo or {}).items():
        limpio[str(cod).strip()] = normalizar_codigo_vendedor(vend) or CODIGO_SIN_ASIGNAR

    nuevo = {}
    for codigo in MAPA_ZONAS.keys():
        if codigo in _HERENCIA_ZONAS:
            origen = _HERENCIA_ZONAS[codigo]
            nuevo[codigo] = limpio.get(origen, CODIGO_SIN_ASIGNAR) if origen else CODIGO_SIN_ASIGNAR
        else:
            nuevo[codigo] = limpio.get(codigo, CODIGO_SIN_ASIGNAR)
    return nuevo

_cache_vinculos = None

def cargar_vinculos_zonas(forzar_recarga=False):
    """Devuelve {codigo_zona: codigo_vendedor}. Se cachea porque se consulta una vez por cliente."""
    global _cache_vinculos
    if _cache_vinculos is not None and not forzar_recarga:
        return _cache_vinculos

    crudo = {}
    if os.path.exists(ARCHIVO_VINCULOS):
        try:
            with open(ARCHIVO_VINCULOS, 'r', encoding='utf-8') as f:
                crudo = json.load(f)
        except:
            crudo = {}

    migrado = _migrar_vinculos(crudo)
    if migrado != crudo:
        try:
            with open(ARCHIVO_VINCULOS, 'w', encoding='utf-8') as f:
                json.dump(migrado, f, indent=4, ensure_ascii=False)
        except: pass

    _cache_vinculos = migrado
    return _cache_vinculos

def guardar_vinculos_zonas(nuevo_mapa):
    global _cache_vinculos
    try:
        limpio = {}
        for cod in MAPA_ZONAS.keys():
            limpio[cod] = normalizar_codigo_vendedor(nuevo_mapa.get(cod, "")) or CODIGO_SIN_ASIGNAR
        with open(ARCHIVO_VINCULOS, 'w', encoding='utf-8') as f:
            json.dump(limpio, f, indent=4, ensure_ascii=False)
        _cache_vinculos = limpio
        return True
    except: return False

# ==========================================
# DETECCIÓN DE ZONA
# ==========================================
def _buscar_codigo_zona(texto_norm):
    """Primer código de 3 cifras del texto que exista en el catálogo (o que sea legado/especial)."""
    for m in _PATRON_CODIGO_ZONA.finditer(texto_norm):
        cod = m.group(1)
        if cod in CODIGOS_ZONA_VALIDOS:
            return cod, "codigo"
        if cod in ZONAS_LEGADO:
            return ZONAS_LEGADO[cod], "codigo_legado"
        if cod in ZONAS_ESPECIALES:
            return cod, "codigo_especial"
    return None, None

def _puntuar_por_localidades(texto_norm):
    """Devuelve {codigo: puntaje} según qué localidades del catálogo aparecen en el texto.
    Las frases largas se consumen primero para que no se cuenten dos veces."""
    restante = texto_norm
    puntajes = {}
    encontrados = []
    for frase in _FRASES_ZONA_ORDENADAS:
        patron = _PATRONES_ZONA[frase]
        if not patron.search(restante):
            continue
        restante = patron.sub(lambda m: " " * len(m.group(0)), restante)
        entrada = _INDICE_ZONAS[frase]
        for cod in entrada["codigos"]:
            puntajes[cod] = puntajes.get(cod, 0) + entrada["peso"]
        encontrados.append(frase)
    return puntajes, encontrados

def _mejores_zonas(puntajes):
    if not puntajes: return []
    tope = max(puntajes.values())
    return sorted([c for c, p in puntajes.items() if p == tope])

def _etiqueta_zona(codigo):
    if codigo in MAPA_ZONAS: return f"{codigo} | {MAPA_ZONAS[codigo]}"
    if codigo in ZONAS_ESPECIALES: return f"{codigo} | {ZONAS_ESPECIALES[codigo]}"
    return str(codigo)

def detectar_zona(texto_fila, zona_cruda="", nombre_cliente=""):
    """Motor de zonas. Devuelve dict con codigo, etiqueta, confianza y motivo.

    Orden de confianza:
      1. Código de zona escrito en la celda de Zona/Localidad  -> altísima
      2. Localidad reconocida en la celda de Zona/Localidad    -> alta
      3. Código de zona en cualquier parte de la fila          -> media
      4. Localidad reconocida en cualquier parte de la fila    -> media/baja
      5. Lo que diga la celda cruda, tal cual                  -> sin clasificar

    Al mirar la fila entera se tapa antes el nombre del cliente, así una razón
    social como 'ELENA SRL' o 'DEVOTO S.A.' no se confunde con esos pueblos.
    """
    zona_norm = normalizar(zona_cruda)
    fila_norm = normalizar(texto_fila)
    nombre_norm = normalizar(nombre_cliente)
    if nombre_norm and len(nombre_norm) >= 4 and nombre_norm != zona_norm:
        fila_norm = fila_norm.replace(nombre_norm, " " * len(nombre_norm))

    if zona_norm:
        cod, origen = _buscar_codigo_zona(zona_norm)
        if cod:
            return {"codigo": cod, "etiqueta": _etiqueta_zona(cod), "confianza": "alta",
                    "motivo": f"código en la celda de zona ({origen})"}

        puntajes, encontrados = _puntuar_por_localidades(zona_norm)
        mejores = _mejores_zonas(puntajes)
        if len(mejores) == 1:
            return {"codigo": mejores[0], "etiqueta": _etiqueta_zona(mejores[0]), "confianza": "alta",
                    "motivo": f"localidad en la celda de zona: {', '.join(encontrados[:3])}"}
        if len(mejores) > 1:
            return {"codigo": None, "etiqueta": " o ".join(mejores) + " | REVISAR (" + ", ".join(encontrados[:3]) + ")",
                    "confianza": "ambigua", "motivo": "la localidad pertenece a más de una zona"}

    cod, origen = _buscar_codigo_zona(fila_norm)
    if cod:
        return {"codigo": cod, "etiqueta": _etiqueta_zona(cod), "confianza": "media",
                "motivo": f"código en la fila ({origen})"}

    puntajes, encontrados = _puntuar_por_localidades(fila_norm)
    mejores = _mejores_zonas(puntajes)
    if len(mejores) == 1:
        conf = "media" if puntajes[mejores[0]] >= PESO_ALIAS else "baja"
        return {"codigo": mejores[0], "etiqueta": _etiqueta_zona(mejores[0]), "confianza": conf,
                "motivo": f"localidad en la fila: {', '.join(encontrados[:3])}"}
    if len(mejores) > 1:
        return {"codigo": None, "etiqueta": " o ".join(mejores) + " | REVISAR (" + ", ".join(encontrados[:3]) + ")",
                "confianza": "ambigua", "motivo": "la localidad pertenece a más de una zona"}

    zc = str(zona_cruda).strip()
    if zc and zc.lower() not in ("nan", "none"):
        return {"codigo": None, "etiqueta": zc, "confianza": "sin clasificar", "motivo": "se dejó el texto original"}
    return {"codigo": None, "etiqueta": "Desconocida", "confianza": "nula", "motivo": "sin datos de ubicación"}

def extraer_zona_inteligente(texto_fila, zona_cruda, nombre_cliente=""):
    """Compatibilidad: devuelve sólo la etiqueta '137 | CORDOBA'."""
    return detectar_zona(texto_fila, zona_cruda, nombre_cliente)["etiqueta"]

# ==========================================
# DETECCIÓN DE VENDEDOR
# ==========================================
def detectar_vendedor(texto_crudo="", vendedor_actual="", zona_o_cobrador="", codigo_zona=None, nombre_cliente=""):
    """Motor de vendedores. Devuelve dict con codigo, nombre, confianza y motivo.

    Prioridad (de más confiable a menos):
      1. Código de vendedor en la columna Vendedor.
      2. Nombre del vendedor en la columna Vendedor.
      3. Código o nombre en la columna Zona/Cobrador (sólo si la celda es un código puro).
      4. Vínculo zona -> vendedor configurado en 'Vincular Zonas'.
      5. Nombre del vendedor suelto en la fila, tapando antes las localidades
         y el nombre del cliente (así 'SAN LUIS' o 'JORGE PEREZ SA' no cuentan).
      6. Sin asignar ('0').
    """
    v_crudo = "" if vendedor_actual is None else str(vendedor_actual).strip()
    z_crudo = "" if zona_o_cobrador is None else str(zona_o_cobrador).strip()

    def _resultado(cod, confianza, motivo):
        info = VENDEDORES_CATALOGO.get(cod, {})
        return {"codigo": cod, "nombre": info.get("nombre", ""), "confianza": confianza, "motivo": motivo}

    # Un "0" explícito significa "sin asignar", así que NO corta la búsqueda:
    # se sigue probando con el vínculo de zona antes de darse por vencido.
    def _util(cod):
        return cod is not None and cod != CODIGO_SIN_ASIGNAR

    # 1) Código explícito en la columna Vendedor
    if v_crudo and _celda_es_solo_codigo(v_crudo):
        cod = normalizar_codigo_vendedor(v_crudo)
        if _util(cod):
            return _resultado(cod, "alta", f"código '{v_crudo}' en la columna Vendedor")

    # 2) Nombre en la columna Vendedor
    if v_crudo:
        cod = _buscar_nombre_vendedor(normalizar(v_crudo))
        if _util(cod):
            return _resultado(cod, "alta", f"nombre '{v_crudo}' en la columna Vendedor")
        # Celda con texto y código mezclados: "VEND 44", "Cod.44 - Alan"
        if not _parece_fecha(v_crudo):
            cod = normalizar_codigo_vendedor(re.sub(r'[^0-9/\-]', ' ', v_crudo))
            if _util(cod) and re.search(r'(?<![0-9])[0-9]{1,3}(?![0-9])', v_crudo):
                return _resultado(cod, "media", f"código dentro de la celda Vendedor ('{v_crudo}')")

    # 3) Columna Zona/Cobrador usada como vendedor (sólo si es un código puro)
    _codigos_que_son_zona = CODIGOS_ZONA_VALIDOS | set(ZONAS_LEGADO) | set(ZONAS_ESPECIALES)
    if z_crudo and _celda_es_solo_codigo(z_crudo):
        cod = normalizar_codigo_vendedor(z_crudo)
        # Ojo: si el "código" es en realidad una zona (101-152), no es un vendedor.
        if _util(cod) and z_crudo.strip() not in _codigos_que_son_zona:
            return _resultado(cod, "media", f"código '{z_crudo}' en la columna Zona/Cobrador")
    if z_crudo:
        cod = _buscar_nombre_vendedor(normalizar(z_crudo))
        if _util(cod):
            return _resultado(cod, "media", f"nombre '{z_crudo}' en la columna Zona/Cobrador")

    # 4) Vínculo por zona
    if codigo_zona:
        vinculos = cargar_vinculos_zonas()
        vend_zona = normalizar_codigo_vendedor(vinculos.get(str(codigo_zona), ""))
        if vend_zona is not None and vend_zona != CODIGO_SIN_ASIGNAR:
            return _resultado(vend_zona, "media", f"vínculo de la zona {codigo_zona}")

    # 5) Nombre suelto en el resto de la fila
    texto_norm = normalizar(texto_crudo)
    if texto_norm:
        nombre_norm = normalizar(nombre_cliente)
        if nombre_norm and len(nombre_norm) >= 4:
            texto_norm = texto_norm.replace(nombre_norm, " " * len(nombre_norm))
        texto_norm = _tapar_geografia(texto_norm)
        cod = _buscar_nombre_vendedor(texto_norm)
        if _util(cod):
            return _resultado(cod, "baja", "nombre del vendedor encontrado en la fila")

    return _resultado(CODIGO_SIN_ASIGNAR, "nula", "no se pudo identificar al vendedor")

def extraer_vendedor_inteligente(texto_crudo, vendedor_actual, zona_o_cobrador, codigo_zona=None, nombre_cliente=""):
    """Compatibilidad: devuelve sólo el código canónico."""
    return detectar_vendedor(texto_crudo, vendedor_actual, zona_o_cobrador, codigo_zona, nombre_cliente)["codigo"]

# ==========================================
# LÓGICA DE LIMPIEZA Y EXTRACCIÓN
# ==========================================
def separar_telefonos(texto_crudo):
    if pd.isna(texto_crudo): return []
    texto = str(texto_crudo).strip()

    # 1. Eliminar CUITs y Fechas típicas para evitar falsos positivos
    texto = re.sub(r'\b\d{2}-\d{8}-\d{1}\b', ' ', texto)
    texto = re.sub(r'\b\d{2}/\d{2}/\d{4}\b', ' ', texto)

    # 2. Reemplazar letras y símbolos "duros" (como /, \, |, *, _, comas) por un delimitador único "|"
    # Mantenemos los números, espacios, guiones (-) y puntos (.) que suelen usarse dentro de un mismo número.
    texto_separado = re.sub(r'[^\d\s\-\.]', '|', texto)

    telefonos_limpios = []

    # 3. Analizar cada bloque separado por nuestro delimitador "|"
    for bloque in texto_separado.split('|'):
        num_puro = ''.join(filter(str.isdigit, bloque))

        if num_puro.startswith("000") or num_puro == "":
            continue

        # Si el bloque ya tiene la longitud ideal de un celular, lo guardamos directo
        if 8 <= len(num_puro) <= 15:
            telefonos_limpios.append(num_puro)

        # Si tiene MÁS de 15 dígitos, es probable que haya dos o más teléfonos distintos pegados
        # por guiones, puntos o espacios (ej: "1145678901 - 1145678902" o "1145678901-1145678902")
        elif len(num_puro) > 15:
            # Sub-dividimos este bloque largo aislando los guiones, puntos o espacios
            sub_bloques = re.split(r'[\s\-\.]+', bloque.strip())
            for sub in sub_bloques:
                sub_puro = ''.join(filter(str.isdigit, sub))
                if 8 <= len(sub_puro) <= 15 and not sub_puro.startswith("000"):
                    telefonos_limpios.append(sub_puro)

    # 4. Retornar la lista eliminando duplicados pero manteniendo el orden de aparición
    vistos = set()
    return [x for x in telefonos_limpios if not (x in vistos or vistos.add(x))]

def estandarizar_columnas(df):
    cols_str = [str(c).lower().strip() for c in df.columns]
    mapa = {}
    asignados = set()
    for original, minuscula in zip(df.columns, cols_str):
        nuevo_nombre = None
        if any(p in minuscula for p in ['vend', 'corredor', 'rep', 'agente']): nuevo_nombre = 'Vendedor'
        elif any(p in minuscula for p in ['zona', 'locali', 'ciudad', 'ubic', 'direc', 'cobr', 'cobrador']): nuevo_nombre = 'Zona_Cruda'
        elif any(p in minuscula for p in ['tel', 'cel', 'móvil', 'movil', 'contacto']): nuevo_nombre = 'Telefonos_Raw'
        elif any(p in minuscula for p in ['cód', 'cod', 'nro', 'código', 'id']): nuevo_nombre = 'Numero_Cliente'
        elif any(p in minuscula for p in ['nombre', 'cliente', 'razon', 'razón', 'social']): nuevo_nombre = 'Nombre'
        if nuevo_nombre and nuevo_nombre not in asignados:
            mapa[original] = nuevo_nombre
            asignados.add(nuevo_nombre)
    df = df.rename(columns=mapa)
    df = df.loc[:, ~df.columns.duplicated()].copy()
    return df

# --- Encabezados basura de los listados del ERP (no son fila de títulos) ---
_ENCABEZADOS_BASURA = ("-zzzz", "-999", "z.fiscal", "ordenado por")

def _adivinar_columna_nombre(df):
    """Cuando la planilla no trae títulos reconocibles, busca cuál columna tiene
    los nombres de los clientes: la que más texto largo tenga, descartando las
    que son localidades, teléfonos o códigos."""
    ya_asignadas = {'Nombre', 'Numero_Cliente', 'Zona_Cruda', 'Telefonos_Raw', 'Vendedor', 'Row_String'}
    mejor, mejor_ratio = None, 0.0
    for col in df.columns:
        if col in ya_asignadas: continue
        serie = df[col].dropna().astype(str).str.strip()
        serie = serie[(serie != "") & (serie.str.lower() != "nan")]
        if len(serie) < 3: continue
        muestra = serie.head(80)
        buenos = 0
        for v in muestra:
            if len(v) < 5 or not _hay_letras(v): continue
            if detectar_zona("", v)["codigo"] is not None: continue   # es una localidad, no un nombre
            buenos += 1
        ratio = buenos / len(muestra)
        if ratio > mejor_ratio and ratio >= 0.5:
            mejor, mejor_ratio = col, ratio
    return mejor

def procesar_un_archivo(ruta):
    try:
        ext = os.path.splitext(ruta)[1].lower()
        dfs_to_process = []

        # --- MOTOR BLINDADO DE LECTURA UNIVERSAL ---
        try:
            if ext in ['.csv', '.txt', '.tsv']:
                try:
                    dfs_to_process = [pd.read_csv(ruta, dtype=str, header=None, encoding='utf-8', on_bad_lines='skip', sep=None, engine='python')]
                except:
                    dfs_to_process = [pd.read_csv(ruta, dtype=str, header=None, encoding='latin-1', on_bad_lines='skip', sep=None, engine='python')]
            elif ext == '.xls':
                try:
                    xls = pd.ExcelFile(ruta, engine='xlrd')
                    dfs_to_process = [pd.read_excel(xls, sheet_name=s, dtype=str, header=None) for s in xls.sheet_names]
                except:
                    dfs_to_process = [pd.read_csv(ruta, dtype=str, header=None, encoding='latin-1', on_bad_lines='skip', sep=None, engine='python')]
            elif ext in ['.xlsx', '.xlsm']:
                try:
                    xls = pd.ExcelFile(ruta, engine='openpyxl')
                    dfs_to_process = [pd.read_excel(xls, sheet_name=s, dtype=str, header=None) for s in xls.sheet_names]
                except:
                    dfs_to_process = [pd.read_excel(ruta, dtype=str, header=None)]
            elif ext == '.xlsb':
                xls = pd.ExcelFile(ruta, engine='pyxlsb')
                dfs_to_process = [pd.read_excel(xls, sheet_name=s, dtype=str, header=None) for s in xls.sheet_names]
            elif ext == '.ods':
                xls = pd.ExcelFile(ruta, engine='odf')
                dfs_to_process = [pd.read_excel(xls, sheet_name=s, dtype=str, header=None) for s in xls.sheet_names]
            else:
                try:
                    dfs_to_process = [pd.read_excel(ruta, dtype=str, header=None)]
                except:
                    dfs_to_process = [pd.read_csv(ruta, dtype=str, header=None, encoding='latin-1', on_bad_lines='skip', sep=None, engine='python')]
        except Exception as inner_e:
            print(f"Error crítico abriendo archivo {ruta}: {inner_e}")
            try:
                dfs_to_process = [pd.read_csv(ruta, dtype=str, header=None, encoding='latin-1', on_bad_lines='skip', sep=None, engine='python')]
            except:
                return pd.DataFrame(), 0

        df_agrupado_total = []
        total_filas = 0

        for df_temp in dfs_to_process:
            if df_temp.empty: continue

            filas_muestra = [list(r) for _, r in df_temp.head(40).fillna("").astype(str).iterrows()]
            fila_enc, _pts = detectar_fila_encabezado(filas_muestra)

            if fila_enc:
                df_temp.columns = df_temp.iloc[fila_enc - 1].fillna(pd.Series([f"Col_{i}" for i in range(len(df_temp.columns))])).astype(str)
                df_temp = df_temp.iloc[fila_enc:].reset_index(drop=True)
            else:
                df_temp.columns = [f"Col_{i}" for i in range(len(df_temp.columns))]

            df_temp = df_temp.dropna(how='all')
            if df_temp.empty: continue

            df_temp = estandarizar_columnas(df_temp)
            # Planilla sin títulos reconocibles: en vez de descartar el archivo entero,
            # se deduce cuál columna tiene los nombres mirando el contenido.
            if 'Nombre' not in df_temp.columns:
                col_nom = _adivinar_columna_nombre(df_temp)
                if col_nom is None: continue
                df_temp = df_temp.rename(columns={col_nom: 'Nombre'})

            total_filas += len(df_temp)
            df_temp['Row_String'] = df_temp.apply(lambda row: ' | '.join(row.dropna().astype(str)), axis=1)

            for col in ['Nombre', 'Numero_Cliente', 'Zona_Cruda', 'Vendedor']:
                if col not in df_temp.columns: df_temp[col] = ""

            df_temp['Numero_Cliente'] = df_temp['Numero_Cliente'].replace(r'^\s*$', np.nan, regex=True).ffill()
            df_temp['Numero_Cliente'] = np.where(df_temp['Numero_Cliente'].isna(), "SinID_" + df_temp.index.astype(str), df_temp['Numero_Cliente'])

            for col in ['Nombre', 'Vendedor', 'Zona_Cruda']:
                df_temp[col] = df_temp[col].replace([r'^\s*$', 'nan', 'None'], np.nan, regex=True)
                df_temp[col] = df_temp.groupby('Numero_Cliente')[col].transform(lambda x: x.ffill().bfill())
                df_temp[col] = df_temp[col].fillna("")

            text_agg = df_temp.groupby('Numero_Cliente')['Row_String'].apply(lambda x: ' | '.join(x.astype(str))).reset_index()
            df_agrupado = df_temp.drop_duplicates(subset=['Numero_Cliente']).copy()
            df_agrupado = df_agrupado.drop(columns=['Row_String']).merge(text_agg, on='Numero_Cliente', how='left')
            df_agrupado_total.append(df_agrupado)

        if not df_agrupado_total: return pd.DataFrame(), 0
        df_final_archivo = pd.concat(df_agrupado_total, ignore_index=True)
        return df_final_archivo, total_filas

    except Exception as e:
        print(f"Archivo omitido por error: {ruta} -> {e}")
        return pd.DataFrame(), 0

def procesar_cruce(df_maestro, progress_callback=None):
    try:
        if progress_callback: progress_callback(5, "Estandarizando memoria en bloque...")
        df = df_maestro.copy()
        cargar_vinculos_zonas(forzar_recarga=True)   # una sola lectura del JSON para todo el cruce
        df['Clave_Agrupacion'] = df['Numero_Cliente'].replace("", np.nan)
        df['Clave_Agrupacion'] = np.where(df['Clave_Agrupacion'].isna(), df['Nombre'].astype(str) + "_" + df.index.astype(str), df['Clave_Agrupacion'])

        if progress_callback: progress_callback(15, "Agrupando clientes duplicados en alta velocidad...")
        for col in ['Nombre', 'Vendedor', 'Zona_Cruda']:
            df[col] = df[col].replace([r'^\s*$', 'nan', 'None'], np.nan, regex=True)
            df[col] = df.groupby('Clave_Agrupacion')[col].transform(lambda x: x.ffill().bfill()).fillna("")

        text_agg = df.groupby('Clave_Agrupacion')['Row_String'].apply(lambda x: ' | '.join(x.astype(str))).reset_index()
        df_agrupado = df.drop_duplicates(subset=['Clave_Agrupacion']).copy()
        df_agrupado = df_agrupado.drop(columns=['Row_String']).merge(text_agg, on='Clave_Agrupacion', how='left')

        datos_procesados = []
        total_filas = len(df_agrupado)
        paso_progreso = max(1, total_filas // 20)

        for idx, (_, row) in enumerate(df_agrupado.iterrows()):
            if progress_callback and total_filas > 0:
                if idx % paso_progreso == 0:
                    progress_callback(15 + int((idx / total_filas) * 85), f"Procesando cliente {idx} de {total_filas}...")

            n = str(row['Nombre']).strip()
            c = str(row['Numero_Cliente']).strip()
            v = str(row['Vendedor']).strip()
            zona_o_cobr = str(row.get('Zona_Cruda', '')).strip()
            texto_total = str(row['Row_String'])

            telefonos_encontrados = separar_telefonos(texto_total)
            zona = detectar_zona(texto_total, zona_o_cobr, n)
            vend = detectar_vendedor(texto_total, v, zona_o_cobr, zona["codigo"], n)

            # --- NUEVA REGLA ULTRA PERMISIVA ---
            # Solo descartamos la fila si literal no tiene nombre ni teléfonos (fila 100% vacía)
            n_low = n.lower()
            if (n == "" or n_low == "nan" or n_low == "cliente sin nombre") and len(telefonos_encontrados) == 0:
                continue

            # Limpiamos basura del sistema de gestión (ERP)
            if "fecha:" in n_low or "hoja:" in n_low or "wood tools" in n_low or "ordenado por" in n_low:
                continue
            if "clientes habilitados" in n_low or "cód." in n_low:
                continue

            registro = {
                'Nombre': n if n not in ["", "nan"] else "Cliente Sin Nombre",
                'Código de cliente': c if not c.startswith("SinID_") else "",
                'Zona del cliente': zona["etiqueta"],
                'Vendedor': vend["codigo"],
                'Primer número': telefonos_encontrados[0] if len(telefonos_encontrados) > 0 else "",
                'Segundo número': telefonos_encontrados[1] if len(telefonos_encontrados) > 1 else "",
                'Tercer número': telefonos_encontrados[2] if len(telefonos_encontrados) > 2 else "",
                'Cuarto número': telefonos_encontrados[3] if len(telefonos_encontrados) > 3 else "",
                'Quinto número': telefonos_encontrados[4] if len(telefonos_encontrados) > 4 else ""
            }
            datos_procesados.append(registro)

        if progress_callback: progress_callback(100, "¡Cruce finalizado!")
        return pd.DataFrame(datos_procesados)
    except Exception as e:
        raise RuntimeError(f"Falla en el motor de cruce: {str(e)}")

def guardar_excel(df_final, ruta_guardar):
    df_final.to_excel(ruta_guardar, index=False)

# ==========================================
# ACTUALIZADOR QUIRÚRGICO DE UBICACIONES
#
# Abre un Excel ya armado y reescribe ÚNICAMENTE las celdas de ubicación.
# No toca rellenos, negritas, bordes, anchos, filtros, fórmulas ni el resto
# de las columnas: se le cambia el .value a la celda y nada más.
# ==========================================
FORMATOS_UBICACION = {
    "completo": "Código + nombre  (ej: 137 | CORDOBA)",
    "codigo":   "Sólo el código   (ej: 137)",
    "nombre":   "Sólo el nombre   (ej: CORDOBA)",
}

# Encabezados candidatos a ser la columna de ubicación, del más al menos probable.
_PISTAS_COL_ZONA = [
    ("zona del cliente", 100), ("localizacion/zona", 95), ("localización/zona", 95),
    ("zona", 90), ("ubicacion", 85), ("ubicación", 85), ("localidad", 80),
    ("ciudad", 70), ("partido", 65), ("provincia", 40), ("region", 40), ("región", 40),
]
_PISTAS_COL_VEND = [("vendedor", 100), ("vend", 80), ("corredor", 70), ("agente", 60), ("representante", 60)]
_PISTAS_COL_NOMBRE = [("razon social", 100), ("razón social", 100), ("nombre", 90), ("cliente", 70)]

def _elegir_columna(encabezados, pistas):
    mejor_idx, mejor_pts = None, 0
    for i, enc in enumerate(encabezados):
        txt = str(enc or "").lower().strip()
        if not txt: continue
        for pista, pts in pistas:
            if pista in txt and pts > mejor_pts:
                mejor_idx, mejor_pts = i, pts
    return mejor_idx

def _formatear_ubicacion(zona, formato):
    if formato == "codigo":
        return zona["codigo"] if zona["codigo"] else zona["etiqueta"]
    if formato == "nombre":
        if zona["codigo"] and zona["codigo"] in MAPA_ZONAS: return MAPA_ZONAS[zona["codigo"]]
        if zona["codigo"] and zona["codigo"] in ZONAS_ESPECIALES: return ZONAS_ESPECIALES[zona["codigo"]]
        return zona["etiqueta"]
    return zona["etiqueta"]

# --- Detección tolerante de la fila de títulos ---------------------------------
# Se puntúa por CATEGORÍA y por coincidencia parcial, así "Nombre del Cliente",
# "RAZON SOCIAL" o "Localización / Zona" también cuentan como títulos.
_CATEGORIAS_ENCABEZADO = {
    "nombre":    (10, ["nombre", "cliente", "razon social", "razón social", "razon", "razón",
                       "apellido", "denominacion", "denominación", "titular", "descripcion",
                       "descripción", "detalle", "empresa", "comercio", "firma"]),
    "codigo":    (5,  ["codigo", "código", "cod.", "cod ", "cód", "nro", "n°", "nº", "numero",
                       "número", "id", "legajo", "cuenta", "cta", "ficha", "padron", "padrón"]),
    "telefono":  (5,  ["telefono", "teléfono", "tel.", "tel ", "fono", "cel", "movil", "móvil",
                       "contacto", "whatsapp", "wsp", "linea", "línea"]),
    "ubicacion": (6,  ["zona", "localidad", "localizacion", "localización", "ubicacion", "ubicación",
                       "ciudad", "pueblo", "partido", "provincia", "region", "región", "domicilio",
                       "direccion", "dirección", "barrio", "plaza", "sucursal", "destino", "reparto",
                       "ruta", "loc."]),
    "vendedor":  (6,  ["vendedor", "vend.", "vend ", "cobrador", "cobr.", "corredor", "agente",
                       "representante", "preventista", "viajante"]),
    "otros":     (3,  ["saldo", "cuit", "condicion", "condición", "iva", "mail", "e-mail", "email",
                       "rubro", "categoria", "categoría", "lista", "observ", "cp", "cod.postal"]),
}

def _categoria_de_titulo(texto):
    t = str(texto).lower().strip()
    if not t or t == "nan" or len(t) > 40:
        return None, 0
    for cat, (pts, palabras) in _CATEGORIAS_ENCABEZADO.items():
        for p in palabras:
            if p in t:
                return cat, pts
    return None, 0

def detectar_fila_encabezado(filas):
    """Busca la fila de títulos de una planilla. Devuelve (fila_1based, puntaje).

    Pide al menos DOS categorías distintas (ej: un título de nombre + uno de teléfono)
    para no confundir una fila de datos que casualmente diga 'cliente'.
    """
    mejor_fila, mejor_pts = None, 0
    for i, valores in enumerate(filas):
        fila_txt = " ".join(str(v) for v in valores if v is not None).lower()
        if any(bas in fila_txt for bas in _ENCABEZADOS_BASURA):
            continue
        vistas = {}
        no_vacias = 0
        for celda in valores:
            if celda is None or not str(celda).strip(): continue
            no_vacias += 1
            cat, pts = _categoria_de_titulo(celda)
            if cat:
                vistas[cat] = max(vistas.get(cat, 0), pts)
        if no_vacias < 2 or len(vistas) < 2:
            continue
        pts = sum(vistas.values())
        if pts > mejor_pts:
            mejor_fila, mejor_pts = i + 1, pts
    return (mejor_fila, mejor_pts) if mejor_pts >= 10 else (None, 0)

def _hay_letras(valor):
    return bool(re.search(r'[A-Za-zÁÉÍÓÚÑáéíóúñ]', str(valor)))

def _detectar_columna_zona_por_contenido(filas, inicio_datos, n_columnas, max_muestras=120):
    """Cuando no hay títulos (o ninguno dice 'zona'), busca la columna cuyo CONTENIDO
    son ubicaciones: la que tenga más localidades/códigos de zona reconocibles."""
    aciertos = [0] * n_columnas
    mirados = [0] * n_columnas
    for fila in filas[inicio_datos:inicio_datos + max_muestras]:
        for c in range(min(n_columnas, len(fila))):
            v = fila[c]
            if v is None or not str(v).strip(): continue
            mirados[c] += 1
            crudo = str(v).strip()
            if crudo in CODIGOS_ZONA_VALIDOS or crudo in ZONAS_LEGADO:
                aciertos[c] += 1
            elif _hay_letras(crudo) and detectar_zona("", crudo)["codigo"] is not None:
                aciertos[c] += 1
    mejor, mejor_ratio = None, 0.0
    for c in range(n_columnas):
        if mirados[c] < 3: continue
        ratio = aciertos[c] / mirados[c]
        if ratio > mejor_ratio and ratio >= 0.30:
            mejor, mejor_ratio = c, ratio
    return mejor

def _letra_columna(indice):
    letra = ""
    n = indice + 1
    while n > 0:
        n, resto = divmod(n - 1, 26)
        letra = chr(65 + resto) + letra
    return letra

def analizar_excel_para_ubicaciones(ruta, max_filas_muestra=200):
    """Mira el Excel sin modificar nada y devuelve, por hoja: dónde está la fila de
    títulos, cómo se llama cada columna, una muestra de su contenido y cuál sería
    la columna de ubicación a reescribir."""
    from openpyxl import load_workbook
    ext = os.path.splitext(ruta)[1].lower()
    wb = load_workbook(ruta, data_only=True, keep_vba=(ext == '.xlsm'))
    reporte = []
    try:
        for hoja in wb.worksheets:
            tope = min(hoja.max_row or 1, max_filas_muestra)
            filas = [list(f) for f in hoja.iter_rows(min_row=1, max_row=tope, values_only=True)]
            n_col = hoja.max_column or (max((len(f) for f in filas), default=0))
            if n_col == 0:
                continue

            fila_enc, _pts = detectar_fila_encabezado(filas)
            encabezados = []
            if fila_enc:
                encabezados = [("" if v is None else str(v).strip()) for v in filas[fila_enc - 1]]
                encabezados += [""] * (n_col - len(encabezados))

            inicio_datos = fila_enc or 0   # índice 0-based de la primera fila de datos

            # Muestra de contenido por columna (para que se pueda elegir a ojo)
            columnas = []
            for c in range(n_col):
                vistos = []
                for fila in filas[inicio_datos:inicio_datos + 40]:
                    if c >= len(fila): continue
                    v = fila[c]
                    if v is None or not str(v).strip(): continue
                    txt = str(v).strip()
                    if txt not in vistos:
                        vistos.append(txt if len(txt) <= 18 else txt[:18] + "…")
                    if len(vistos) >= 3: break
                columnas.append({
                    "indice": c,
                    "letra": _letra_columna(c),
                    "titulo": encabezados[c] if c < len(encabezados) else "",
                    "muestra": ", ".join(vistos),
                })

            col_zona = _elegir_columna(encabezados, _PISTAS_COL_ZONA) if encabezados else None
            if col_zona is None:
                col_zona = _detectar_columna_zona_por_contenido(filas, inicio_datos, n_col)

            reporte.append({
                "hoja": hoja.title,
                "fila_encabezado": fila_enc,
                "encabezados": encabezados,
                "columnas": columnas,
                "col_zona": col_zona,
                "col_vendedor": _elegir_columna(encabezados, _PISTAS_COL_VEND) if encabezados else None,
                "col_nombre": _elegir_columna(encabezados, _PISTAS_COL_NOMBRE) if encabezados else None,
                "filas_totales": hoja.max_row or 0,
                "columnas_totales": n_col,
            })
    finally:
        wb.close()
    return reporte

def actualizar_ubicaciones_excel(ruta_origen, ruta_destino=None, plan=None, formato="completo",
                                 solo_vacias=False, actualizar_vendedor=False,
                                 pisar_formulas=False, progress_callback=None):
    """Reescribe SÓLO las celdas de ubicación del Excel indicado.

    ruta_origen  : Excel a leer (.xlsx / .xlsm).
    ruta_destino : dónde guardar. Si es None, se guarda una copia
                   '<nombre>_ZONAS_ACTUALIZADAS.xlsx' al lado del original.
    plan         : {nombre_hoja: {'fila_encabezado': int, 'col_zona': idx | 'NUEVA' | None,
                                  'col_vendedor': idx|None, 'col_nombre': idx|None}}
                   Si es None se usa la detección automática de analizar_excel_para_ubicaciones().
    formato      : 'completo' | 'codigo' | 'nombre'.
    solo_vacias  : True = sólo completa las celdas de ubicación que hoy están vacías.
    actualizar_vendedor : también reescribe la columna Vendedor (misma lógica quirúrgica).
    pisar_formulas      : False = no toca celdas que contengan fórmulas.

    Conserva rellenos, negritas, bordes, formatos numéricos, anchos, filtros,
    validaciones, hojas ocultas y todas las demás columnas.
    """
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from copy import copy as copiar_estilo

    ext = os.path.splitext(ruta_origen)[1].lower()
    if ext not in ('.xlsx', '.xlsm'):
        raise ValueError("Para actualizar conservando el formato el archivo tiene que ser .xlsx o .xlsm "
                         f"(se recibió '{ext}'). Convertilo primero desde Excel.")

    if ruta_destino is None:
        base, _ = os.path.splitext(ruta_origen)
        ruta_destino = f"{base}_ZONAS_ACTUALIZADAS{ext}"

    if progress_callback: progress_callback(3, "Abriendo el archivo...")

    # Dos lecturas: una con los valores ya calculados (para leer bien las fórmulas)
    # y otra con las fórmulas intactas (es la que se guarda).
    try:
        wb_valores = load_workbook(ruta_origen, data_only=True, keep_vba=(ext == '.xlsm'))
    except Exception:
        wb_valores = None
    wb = load_workbook(ruta_origen, data_only=False, keep_vba=(ext == '.xlsm'))

    if plan is None:
        plan = {}
        for info in analizar_excel_para_ubicaciones(ruta_origen):
            plan[info["hoja"]] = {
                "fila_encabezado": info["fila_encabezado"],
                "col_zona": info["col_zona"],
                "col_vendedor": info["col_vendedor"],
                "col_nombre": info["col_nombre"],
            }

    cargar_vinculos_zonas(forzar_recarga=True)

    reporte = {
        "ruta_destino": ruta_destino,
        "hojas": [],
        "actualizadas": 0,
        "sin_cambios": 0,
        "omitidas_formula": 0,
        "omitidas_combinada": 0,
        "sin_ubicacion": 0,
        "vendedores_actualizados": 0,
    }

    hojas_con_plan = [h for h in wb.sheetnames if h in plan and plan[h].get("col_zona") is not None]
    total_hojas = max(1, len(hojas_con_plan))

    for n_hoja, nombre_hoja in enumerate(hojas_con_plan):
        conf = plan[nombre_hoja]
        hoja = wb[nombre_hoja]
        hoja_val = wb_valores[nombre_hoja] if wb_valores is not None and nombre_hoja in wb_valores.sheetnames else None

        # 0 (o None) = la hoja no tiene fila de títulos y los datos arrancan en la fila 1
        fila_encabezado = conf.get("fila_encabezado") or 0
        col_zona = conf.get("col_zona")
        col_vend = conf.get("col_vendedor")
        col_nombre = conf.get("col_nombre")

        # Columna nueva al final (se le copia el estilo del encabezado vecino para que no desentone)
        if col_zona == "NUEVA":
            col_zona = hoja.max_column          # índice 0-based de la columna nueva = max_column actual
            if fila_encabezado >= 1:
                celda_nueva = hoja.cell(row=fila_encabezado, column=col_zona + 1)
                if not isinstance(celda_nueva, MergedCell):
                    vecina = hoja.cell(row=fila_encabezado, column=max(1, col_zona))
                    try:
                        celda_nueva._style = copiar_estilo(vecina._style)
                    except Exception:
                        pass
                    celda_nueva.value = "Zona del cliente"

        detalle_hoja = {"hoja": nombre_hoja, "actualizadas": 0, "sin_cambios": 0,
                        "omitidas_formula": 0, "omitidas_combinada": 0, "sin_ubicacion": 0,
                        "columna": col_zona + 1}

        primera_fila_datos = fila_encabezado + 1
        ultima_fila = hoja.max_row or 0
        total_filas = max(1, ultima_fila - primera_fila_datos + 1)

        for nro_fila in range(primera_fila_datos, ultima_fila + 1):
            if progress_callback and nro_fila % 200 == 0:
                avance = (n_hoja + (nro_fila - primera_fila_datos) / total_filas) / total_hojas
                progress_callback(5 + int(avance * 90), f"{nombre_hoja}: fila {nro_fila} de {ultima_fila}...")

            fila_lectura = hoja_val[nro_fila] if hoja_val is not None else hoja[nro_fila]
            valores = [c.value for c in fila_lectura]
            if not any(v not in (None, "") for v in valores):
                continue

            texto_fila = " | ".join(str(v) for v in valores if v not in (None, ""))
            valor_zona_actual = valores[col_zona] if col_zona < len(valores) else None
            nombre_cliente = str(valores[col_nombre]) if (col_nombre is not None and col_nombre < len(valores) and valores[col_nombre]) else ""

            celda = hoja.cell(row=nro_fila, column=col_zona + 1)
            if isinstance(celda, MergedCell):
                detalle_hoja["omitidas_combinada"] += 1
                continue
            if not pisar_formulas and isinstance(celda.value, str) and celda.value.startswith("="):
                detalle_hoja["omitidas_formula"] += 1
                continue
            if solo_vacias and celda.value not in (None, ""):
                detalle_hoja["sin_cambios"] += 1
                continue

            zona = detectar_zona(texto_fila, valor_zona_actual or "", nombre_cliente)
            if zona["codigo"] is None and zona["confianza"] != "ambigua":
                # No se pudo reconocer la ubicación: la celda se deja exactamente como estaba.
                detalle_hoja["sin_ubicacion"] += 1
                continue

            nuevo_valor = _formatear_ubicacion(zona, formato)
            if str(celda.value or "").strip() == str(nuevo_valor).strip():
                detalle_hoja["sin_cambios"] += 1
            else:
                celda.value = nuevo_valor      # <-- lo único que se toca de la planilla
                detalle_hoja["actualizadas"] += 1

            if actualizar_vendedor and col_vend is not None:
                celda_v = hoja.cell(row=nro_fila, column=col_vend + 1)
                if isinstance(celda_v, MergedCell): continue
                if not pisar_formulas and isinstance(celda_v.value, str) and celda_v.value.startswith("="): continue
                valor_vend_actual = valores[col_vend] if col_vend < len(valores) else ""
                vend = detectar_vendedor(texto_fila, valor_vend_actual or "", valor_zona_actual or "",
                                         zona["codigo"], nombre_cliente)
                # Si no se pudo identificar al vendedor, la celda queda como estaba:
                # nunca se escribe "sin asignar" encima de un dato existente ni de un vacío.
                if vend["codigo"] != CODIGO_SIN_ASIGNAR and str(celda_v.value or "").strip() != vend["codigo"]:
                    celda_v.value = vend["codigo"]
                    reporte["vendedores_actualizados"] += 1

        for clave in ("actualizadas", "sin_cambios", "omitidas_formula", "omitidas_combinada", "sin_ubicacion"):
            reporte[clave] += detalle_hoja[clave]
        reporte["hojas"].append(detalle_hoja)

    if progress_callback: progress_callback(96, "Guardando el archivo...")
    wb.save(ruta_destino)
    wb.close()
    if wb_valores is not None: wb_valores.close()
    if progress_callback: progress_callback(100, "¡Ubicaciones actualizadas!")
    return reporte
